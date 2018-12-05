# encoding=utf8

from openpyxl import load_workbook
import re

import sys
reload(sys)
sys.setdefaultencoding('utf8')

file = load_workbook(filename='advisory-activities.xlsx')
activeSheet = file.active
sheets = file.sheetnames
data = ''

def convert(name):
    for k in name.split('\n'):
        return re.sub(r"[^a-zA-Z0-9]+", ' ', k)

for sheet in sheets:
    headers = []
    line_count = 0
    for row in file[sheet].iter_rows():
        if line_count > 0:
            # data += '*** \n ### New Advisory Activity ### \n *** \n'
            data = ''
        for cellIndex, cell in enumerate(row):
            if cell.column == 'A':
                mdFileName = convert(cell.value)
                mdFileName = mdFileName.replace(' ', '-')
            if cell.value != None:
                if line_count == 0:
                    headers.append('#### ' + cell.value)
                else:
                    data += headers[cellIndex] + '\n'
                    cell.value = cell.value.replace('!&-&!', '–').replace('!*-*!', '—')
                    data += cell.value + '\n'
            with open('md/' + mdFileName + '.md', 'w') as f:
                f.write(data)
        line_count += 1
