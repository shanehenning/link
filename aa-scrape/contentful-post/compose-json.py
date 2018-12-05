from openpyxl import load_workbook
import json

file = load_workbook(filename='advisory-activities-v4-voice-updates-LWP.xlsx')
activeSheet = file.active
sheets = file.sheetnames

md = {
'Class Challenge': ['objective', 'materials', 'themes', 'prep', 'activityInstructions', 'introduction', 'steps', 'reflection', 'additionalSection', 'description'],
'Class Meeting': ['objective', 'themes', 'preparation', 'instructions', 'warm-up', 'discussion', 'reflection', 'description'],
'Service Learning':['objective', 'materials', 'themes', 'preparation', 'projectDescription', 'instructionsHeader', 'investigation', 'planning', 'action', 'reflection', 'demonstration', 'description']
}
allData = []
data = {}
default = ''
def find(string, text):
    if text in string:
        return True
    else:
        return False

def camelCase(string):
    string = ''.join(x for x in cell.value.title() if not x.isspace())
    return string[0].lower() + string[1:]

for sheet in sheets:
    headers = []
    line_count = 0
    for row in file[sheet].iter_rows():
        if sheet == 'Service Learning':
            data = {}
            if line_count > 3:
                break
            for cellIndex, cell in enumerate(row):
                if cell.value != None:
                    if line_count == 0:
                        headers.append(camelCase(cell.value))
                    else:
                        data[headers[cellIndex]] = {}
                        # if headers[cellIndex] in md[sheet]:
                        #     data[headers[cellIndex]]['en-US'] = {}
                        #     data[headers[cellIndex]]['en-US']['content'] = []
                        #     data[headers[cellIndex]]['en-US']['content'].append({"content": [{"value": cell.value}]})
                        # else:
                        data[headers[cellIndex]]['en-US'] = cell.value
            if line_count > 0:
                if data != {}:
                    allData.append(data)
            line_count += 1
        with open('test2.json', 'w') as f:
            f.write(json.dumps(allData))
