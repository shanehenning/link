from openpyxl import load_workbook
import json

workbook = load_workbook(filename='scrape_sample.xlsx')
sheets = workbook.sheetnames
# activitiesWorkbook = load_workbook(filename='advisory-activities.xlsx')

page_metadata = {'page': [], 'pdf': []}
for sheet in sheets:
    line_count = 0
    for row in workbook[sheet].iter_rows(min_row=1):
        if sheet == 'scrape':
            if line_count > 0:
                new_page = {}
                for cell in row:
                    if cell.column == 'B':
                        new_page['title'] = cell.value
                    if cell.column == 'C':
                        new_page['gradeRange'] = cell.value.split(', ')
                    if cell.column == 'D':
                        new_page['product'] = cell.value.split(', ')
                    if cell.column == 'F':
                        new_page['audience'] = cell.value.split(', ')
                page_metadata['page'].append(new_page)
            line_count += 1
        if sheet == 'pdf':
            if line_count > 0:
                new_pdf = {}
                for cell in row:
                    if cell.column == 'A':
                        new_pdf['fileName'] = cell.value
                    if cell.column == 'B':
                        new_pdf['title'] = cell.value
                    if cell.column == 'D':
                        new_pdf['gradeRange'] = cell.value.split(', ')
                    if cell.column == 'G':
                        new_pdf['product'] = cell.value.split(', ')
                    if cell.column == 'H':
                        if str(cell.value).lower() == 'false':
                            new_pdf['writeable'] = False
                        if str(cell.value).lower() == 'true':
                            new_pdf['writeable'] = True
                    if cell.column == 'I':
                        new_pdf['audience'] = cell.value.split(', ')
                    if cell.column == 'J':
                        if str(cell.value).lower() == 'false':
                            new_pdf['resourceLibrary'] = False
                        if str(cell.value).lower() == 'true':
                            new_pdf['resourceLibrary'] = True
                page_metadata['pdf'].append(new_pdf)
            line_count += 1


with open('metadata.json', 'w') as f:
    f.write(json.dumps(page_metadata))
