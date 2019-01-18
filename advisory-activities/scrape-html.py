from openpyxl import load_workbook
from bs4 import BeautifulSoup
import tomd

indexWorkbook = load_workbook(filename='results.xlsx')
indexFile = indexWorkbook.active
activitiesWorkbook = load_workbook(filename='advisory-activities.xlsx')

line_count = 0



# soup extraction functions
def getText(entry, attr, selector, *optionalSecondElement):
    if optionalSecondElement:
        return soup.find(entry, attrs={attr: selector}).find(optionalSecondElement).text.strip()
    else:
        return soup.find(entry, attrs={attr: selector}).text.strip()

def getList(entry, attr, selector, element):
        list = soup.find(entry, attrs={attr: selector}).find_all(element)
        list_md = ''
        for idx, val in enumerate(list):
            if element == 'p':
                if val.get('class') == [u'counter']:
                    list_md += '- ' + val.text.strip() + '\n'
                else:
                    list_md += '\n' + tomd.convert(str(val)).strip() + '\n'
            else:
                list_md += '- ' + val.text.strip() + '\n'
        return list_md

def getThemes(entry, attr, selector, secondElement):
    list_items = soup.find(entry, attrs={attr: selector}).find_all(secondElement)
    themes_md = ''
    for idx, val in enumerate(list_items):
        img = val.find('img')
        text = val.find('h4')
        themes_md += '![' + img.get('alt') + ']' + '(' + img.get('src') + ')'
        themes_md += text.text.strip()
        themes_md += '\n \n'
    return themes_md

def getMarkdown(entry, attr, selector, *args, **kwargs):
    if args or kwargs:
        if args and kwargs:
            snippet = soup.find_all(entry, attrs={attr, selector})[args[0]].find(kwargs['opt_entry'], attrs={kwargs['opt_attr'], kwargs['opt_selector']})
        elif kwargs and len(args) == 0:
            snippet = soup.find(entry, attrs={attr, selector}).find(kwargs['opt_entry'], attrs={kwargs['opt_attr'], kwargs['opt_selector']})
        uls = snippet.find('ul')
        if uls != None:
            snippet.ul.append(soup.new_tag('p'))
        snippet = tomd.convert(str(snippet)).replace('-', '\n- ').strip()
        return snippet
    return tomd.convert(str(soup.find(entry, attrs={attr, selector})))
# end soup extraction functions

def convertText(text):
    text = text.encode('ascii', 'ignore')
    return text.replace(u"\u2018", "'").replace(u"\u2019", "'").replace(u'\u201C', '"').replace(u'\u201D', '"').replace(u'\u2014', '--').strip()

# scrape function
for row in indexFile.iter_rows(min_row=1):
    # print('col: ', col)
    for cell in row:
        print('cell: ', cell)
        if cell.value == None:
            break
        cell.value = convertText(cell.value)
        soup = BeautifulSoup(str(cell.value), 'html.parser')
        category = getText('h5', 'class', 'text-activity-category')
        categoryType = category.split(':')[0].lower()
        if categoryType == 'class challenge':
            name = getText('h1', 'class', 'text-activity-name')
            objective = getText('div', 'class', 'activity-details', 'p')
            materials_md = getList('div', 'class', 'activity-details-all', 'li')
            themes_md = getThemes('ul', 'class', 'activity-themes', 'li')
            preparation_md = getList('div', 'id', 'collapsePreparation', 'p')
            instructions_header = getText('h5', 'class', 'text-instructions')
            introduction_md = getMarkdown('article', 'class', 'panel-introduction', opt_entry='div', opt_attr='class', opt_selector='panel-content')
            steps_md = getList('article', 'class', 'panel-steps', 'p')
            reflection_md = getList('article', 'class', 'panel-reflection', 'p')

        elif categoryType == 'class meeting':
            name = getText('h1', 'class', 'text-activity-name')
            objective = getText('div', 'class', 'activity-details', 'p')
            themes_md = getThemes('ul', 'class', 'activity-themes', 'li')
            preparation_md = getList('div', 'id', 'collapsePreparation', 'p')
            instructions_header = getText('h5', 'class', 'text-instructions')
            warm_up_md = getMarkdown('article', 'class', 'panel-introduction', opt_entry='div', opt_attr='class', opt_selector='panel-content')
            discussion_md = getMarkdown('article', 'class', 'panel-steps', opt_entry='div', opt_attr='class', opt_selector='panel-content')
            reflection_md = getMarkdown('article', 'class', 'panel-reflection', opt_entry='div', opt_attr='class', opt_selector='panel-content')

        elif categoryType == 'service learning':
            name = getText('h1', 'class', 'text-activity-name')
            objective = getList('div', 'class', 'activity-details-all', 'li')
            # materials = getText('div', 'class', 'activity-details-all', 'p')
            # materials = getMarkdown('div', 'class', 'activity-details-all', 1, opt_entry='div', opt_attr='class', opt_selector='activity-details')
            materials_md = soup.find('div', attrs={'class', 'activity-details-all'}).find_all('div', attrs={'class', 'activity-details'})[1]
            materials_uls = materials_md.find('ul')
            if materials_uls != None:
                materials_md.ul.append(soup.new_tag('p'))
            materials = tomd.convert(str(materials_md)).replace('-', '\n- ').strip()
            themes_md = getThemes('ul', 'class', 'activity-themes', 'li')
            preparation_md = getList('div', 'id', 'collapsePreparation', 'p')
            project_description = getMarkdown('div', 'class', 'activity-project-description')
            instructions_header = getText('h5', 'class', 'text-instructions')
            investigation_md = getMarkdown('article', 'class', 'panel', 0, opt_entry='div', opt_attr='class', opt_selector='panel-content')
            planning_md = getMarkdown('article', 'class', 'panel', 1, opt_entry='div', opt_attr='class', opt_selector='panel-content')
            action = getMarkdown('article', 'class', 'panel', 2, opt_entry='div', opt_attr='class', opt_selector='panel-content')
            reflection = getMarkdown('article', 'class', 'panel', 3, opt_entry='div', opt_attr='class', opt_selector='panel-content')
            demonstration = getMarkdown('article', 'class', 'panel', 4, opt_entry='div', opt_attr='class', opt_selector='panel-content')

        else:
            print 'not a class challenge or class meeting or service learning'
        line_count += 1
        if categoryType == 'class challenge':
            activitiesWorkbook[u'Class Challenge'].append([name, category, objective, materials_md, themes_md, preparation_md, instructions_header, introduction_md, steps_md, reflection_md])
        elif categoryType == 'class meeting':
            activitiesWorkbook[u'Class Meeting'].append([name, category, objective, themes_md, preparation_md, instructions_header, warm_up_md, discussion_md, reflection_md])
        elif categoryType == 'service learning':
            activitiesWorkbook[u'Service Learning'].append([name, category, objective, materials, themes_md, preparation_md, project_description, instructions_header, investigation_md, planning_md, action, reflection, demonstration])
        else:
            print 'categoryType: ' + categoryType
activitiesWorkbook.save('advisory-activities.xlsx')
# end scrape function


print 'done'
