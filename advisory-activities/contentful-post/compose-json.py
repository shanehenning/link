# encoding=utf8

from openpyxl import load_workbook
import json
import re

import sys
reload(sys)
sys.setdefaultencoding('utf8')

file = load_workbook(filename='advisory-activities-v5-voice-updates-LWP.xlsx')
sheets = file.sheetnames

allData = []
data = {}

def camelCase(string):
    string = ''.join(x for x in cell.value.title() if not x.isspace())
    return string[0].lower() + string[1:]

for sheet in sheets:
    headers = []
    line_count = 0
    for row in file[sheet].iter_rows():
        if sheet == 'Service Learning':
            data = {}

            to limit the amount of activities processed
            if line_count > 3:
                break

            for cellIndex, cell in enumerate(row):
                if cell.value != None:
                    if line_count == 0:
                        headers.append(camelCase(cell.value))
                    else:
                        cell.value = cell.value.replace('!&-&!', '–').replace('!%-%!', '—')
                        data[headers[cellIndex]] = {}
                        data[headers[cellIndex]]['en-US'] = cell.value
            if line_count > 0:
                if data != {}:
                    allData.append(data)
            line_count += 1
        with open('test-output.json', 'w') as f:
            f.write(json.dumps(allData))
